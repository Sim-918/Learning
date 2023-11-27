import random
import pandas as pd
from openpyxl import load_workbook

wb=load_workbook('".\단어 시험_data\day1.xlsx"')
data=wb.active
print(data['A1'].value)

# data=pd.read_excel(".\단어 시험_data\day1.xlsx",index_col=0)
# print(data)

# for i in ws.iter_rows:
    # print(i)

# word=[['a','에이'],['b','비'],['c','씨'],['d','디'],['e','이'],['f','에프']]
# sunji_list=word*2
# # print(sunji_list)

# cnt=0

# for i in range(len(word)):
#     test=[word[i][1]]
#     print(f"문제: ",word[i][0])
#     tmp=sunji_list[i+1:i+4]
#     for j in tmp:
#         test.append(j[1])
#     random.shuffle(test)
#     print(test)
#     user_input=int(input())
#     if word[i][1]==test[user_input-1]:
#         cnt+=1
#         print("correct")
#     else:
#         print("you dead")
# print(f"문항수 : ",len(word))
# print(f"맞춘 수: ",cnt)
